"""
config_manager.py
역할: column_config.yaml 로드/저장/생성 + Excel 파일 대상 사전 검증.

저장 경로 규칙:
    - EXE 배포 환경 (PyInstaller frozen): sys.executable 옆 폴더
      → sys._MEIPASS 는 임시 폴더이므로 재시작 시 삭제됨, 절대 사용 금지
    - 개발 환경: 프로젝트 루트 (Path(__file__).parents[3])
"""
from __future__ import annotations

import copy
import difflib
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml

# ── 설정 파일 경로 ─────────────────────────────────────────────────────────────
# PyInstaller frozen 환경에서는 sys.executable(EXE 파일) 옆에 저장한다.
# sys._MEIPASS 는 프로그램 종료 시 삭제되는 임시 폴더이므로 사용하면 안 된다.
if getattr(sys, "frozen", False):
    _CONFIG_DIR = Path(sys.executable).parent
else:
    _CONFIG_DIR = Path(__file__).parents[3]

_CONFIG_PATH = _CONFIG_DIR / "column_config.yaml"

# ── 기본 설정값 — data_loader.COLUMN_MAP + _SHEET_* 상수의 단일 원본 ──────────
_DEFAULT_CONFIG: dict[str, Any] = {
    "sheets": {
        "transaction": "실제발생사업비",
        "account":     "계정정보",
        "ccm":         "Cost Center Master",
        "output":      "출력>",
    },
    "columns": {
        # 실제발생사업비 시트
        "원가요소":        "원가요소",
        "코스트센터":      "코스트 센터",
        # 계정정보 시트
        "계정번호":        "계정번호",
        "계정명":          "계정명",
        "계정그룹ID":      "계정그룹ID",
        "계정그룹명":      "계정그룹명",
        "직간접구분_계정": "직/간접비",
        "대상정의":        "대상정의 v3.0_0415",
        "범위":            "사용 부서",
        "지급대상":        "비용 지급 범위",
        "산출기준":        "산출기준",
        # Cost Center Master 시트
        "cc_code":        "Cost Center Code",
        "cc_name":        "Cost Center name",
        "직간접구분":      "변경: 직접/공통 구분",
        # 출력 시트
        "출력전표":        "출력전표",
        "귀속_주관부서":   "귀속_주관부서",
        "귀속_사용부서":   "귀속_사용부서",
    },
    "ccm_rename": {
        "Code":   "본부_Code",
        "Code.1": "팀_Code",
    },
    "amount_cols": [
        "Sum of DA_P", "Sum of DA_N", "Sum of DM", "Sum of DC",
        "Sum of DI",  "Sum of DP",   "Sum of IM",  "Sum of II",
        "Sum of IO",  "Sum of IA",
    ],
    "nature_cols": [
        "계약비", "유지비", "손해조사비", "투자관리비", "기타사업비",
    ],
    "col_team": "팀",
    "output_labels": {
        "코드": "출력전표",
        "주관": "주관부서",
        "사용": "사용부서",
    },
}

# ── 시트별 검증 대상 컬럼 키 목록 ─────────────────────────────────────────────
# JOIN 키 및 집계 필수 컬럼만 — 모든 컬럼 검증은 선택적 컬럼 누락을 오탐으로 처리함
_SHEET_REQUIRED_COLS: dict[str, list[str]] = {
    "transaction": ["원가요소", "코스트센터"],
    "account":     ["계정번호", "대상정의", "직간접구분_계정"],
    "ccm":         ["cc_code", "cc_name", "직간접구분"],
    "output":      ["출력전표"],
}

# ── 시트 논리명 → 표시용 한국어 레이블 ───────────────────────────────────────
SHEET_LABELS: dict[str, str] = {
    "transaction": "실제발생사업비",
    "account":     "계정정보",
    "ccm":         "Cost Center Master",
    "output":      "출력전표 시트",
}

# ── 컬럼 논리명 → 표시용 한국어 설명 ─────────────────────────────────────────
_COL_LABELS: dict[str, str] = {
    "원가요소":        "원가요소 (필터·JOIN 키)",
    "코스트센터":      "코스트 센터 (JOIN 키)",
    "계정번호":        "계정번호 (JOIN 키)",
    "대상정의":        "대상정의 (보고서 본문)",
    "직간접구분_계정": "직/간접비 (분류 기준)",
    "cc_code":        "Cost Center Code (JOIN 키)",
    "cc_name":        "Cost Center name (표시용)",
    "직간접구분":      "직간접 구분 (CCM 분류)",
    "출력전표":        "출력전표 번호",
}


# ════════════════════════════════════════════════════════════════════════════
#  Mismatch 데이터 클래스
# ════════════════════════════════════════════════════════════════════════════

@dataclass
class Mismatch:
    """설정 파일과 실제 Excel 파일 간 불일치 정보를 담는 데이터 클래스."""

    sheet_key: str           # "account" 등 내부 논리 시트 키
    logical_name: str        # "대상정의" 등 columns 딕셔너리 키
    expected: str            # 설정 파일의 현재 값
    actual_cols: list[str]   # 해당 시트의 실제 값 목록 (드롭다운용)
    fuzzy_guess: str | None  # Fuzzy 매칭 추천값 (없으면 None)
    is_sheet_mismatch: bool  # True = 시트명 불일치, False = 컬럼명 불일치
    display_label: str = field(default="", init=False)

    def __post_init__(self) -> None:
        if self.is_sheet_mismatch:
            label = SHEET_LABELS.get(self.sheet_key, self.sheet_key)
            self.display_label = f"{label} (시트명)"
        else:
            self.display_label = _COL_LABELS.get(self.logical_name, self.logical_name)


@dataclass
class ColWarning:
    """선택적 컬럼(AMOUNT_COLS/NATURE_COLS) 누락 경고."""

    sheet_key: str           # "transaction"
    category: str            # "amount_cols" | "nature_cols"
    missing_cols: list[str]  # 누락된 컬럼명 목록
    found_cols: list[str]    # 해당 시트의 실제 컬럼 목록 (참고용)
    display_label: str = field(default="", init=False)

    def __post_init__(self) -> None:
        label = SHEET_LABELS.get(self.sheet_key, self.sheet_key)
        cat_label = "금액 컬럼" if self.category == "amount_cols" else "성격별 분류 컬럼"
        self.display_label = f"{label} — {cat_label} 누락"




# ════════════════════════════════════════════════════════════════════════════
#  Public API
# ════════════════════════════════════════════════════════════════════════════

def load_config() -> dict[str, Any]:
    """column_config.yaml을 로드하고 기본값과 deep-merge하여 반환한다.

    파일이 없으면 주석 포함 YAML을 자동 생성한 뒤 기본값을 반환한다.
    파일이 있으나 파싱 실패 또는 비정상 구조 시 기본값으로 폴백한다.
    누락된 키는 기본값으로 채운다.
    """
    if not _CONFIG_PATH.exists():
        generate_default_config_file()
        return copy.deepcopy(_DEFAULT_CONFIG)

    try:
        raw = _CONFIG_PATH.read_text(encoding="utf-8")
        loaded = yaml.safe_load(raw)
        if not isinstance(loaded, dict):
            return copy.deepcopy(_DEFAULT_CONFIG)
        return _deep_merge(_DEFAULT_CONFIG, loaded)
    except Exception:
        return copy.deepcopy(_DEFAULT_CONFIG)


def save_config(config: dict[str, Any]) -> None:
    """설정을 column_config.yaml에 한국어 주석 헤더와 함께 저장한다.

    OS 오류(쓰기 권한 부족 등)는 조용히 무시한다.
    """
    try:
        body = yaml.dump(
            config,
            allow_unicode=True,
            default_flow_style=False,
            indent=2,
            sort_keys=False,
        )
        _CONFIG_PATH.write_text(_YAML_HEADER + body, encoding="utf-8")
    except OSError:
        pass


def generate_default_config_file() -> None:
    """기본값으로 column_config.yaml을 생성한다.

    각 항목에 한국어 주석을 포함하여 비개발자도 이해할 수 있는 형태로 작성한다.
    파일이 이미 존재하면 덮어쓰지 않는다.
    """
    if _CONFIG_PATH.exists():
        return
    try:
        _CONFIG_PATH.write_text(_YAML_HEADER + _ANNOTATED_YAML_BODY, encoding="utf-8")
    except OSError:
        pass


def validate_against_excel(wb_path: str, config: dict[str, Any]) -> list[Mismatch]:
    """Excel 파일을 openpyxl read_only로 열고 설정값과 비교한다.

    시트명 불일치 → 컬럼명 불일치 순으로 검사한다.
    시트 자체가 없으면 해당 시트의 컬럼 검증은 건너뛴다.

    Args:
        wb_path: 입력 Excel 파일 경로
        config:  load_config()의 반환값

    Returns:
        불일치 항목의 Mismatch 리스트. 모두 일치 시 빈 리스트.

    Raises:
        ValueError: Excel 파일 자체를 열 수 없을 때
    """
    import openpyxl  # noqa: PLC0415

    try:
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Excel 파일을 열 수 없습니다: {exc}") from exc

    actual_sheet_names: list[str] = list(wb.sheetnames)
    sheets_cfg = config.get("sheets", {})
    cols_cfg = config.get("columns", {})
    mismatches: list[Mismatch] = []

    try:
        for sheet_key, required_col_keys in _SHEET_REQUIRED_COLS.items():
            expected_sheet = sheets_cfg.get(sheet_key, "")

            # ── 시트명 검증 ─────────────────────────────────────────────────
            if expected_sheet not in actual_sheet_names:
                fuzzy = _fuzzy_match(expected_sheet, actual_sheet_names)
                mismatches.append(Mismatch(
                    sheet_key=sheet_key,
                    logical_name=sheet_key,
                    expected=expected_sheet,
                    actual_cols=actual_sheet_names,
                    fuzzy_guess=fuzzy,
                    is_sheet_mismatch=True,
                ))
                continue  # 시트가 없으면 컬럼 검증 불가

            # ── 컬럼명 검증 ─────────────────────────────────────────────────
            actual_cols = _get_sheet_columns(wb, expected_sheet)

            for col_key in required_col_keys:
                expected_col = cols_cfg.get(col_key, "")
                if not expected_col:
                    continue
                if expected_col not in actual_cols:
                    fuzzy = _fuzzy_match(expected_col, actual_cols)
                    mismatches.append(Mismatch(
                        sheet_key=sheet_key,
                        logical_name=col_key,
                        expected=expected_col,
                        actual_cols=actual_cols,
                        fuzzy_guess=fuzzy,
                        is_sheet_mismatch=False,
                    ))
    finally:
        wb.close()

    return mismatches


def validate_optional_cols(wb_path: str, config: dict[str, Any]) -> list[ColWarning]:
    """실제발생사업비 시트에서 AMOUNT_COLS / NATURE_COLS 누락 여부를 확인한다.

    결과가 0 또는 잘못 집계될 수 있는 선택적 컬럼 부재를 경고로 알린다.
    이 경고는 처리를 중단하지 않고 사용자에게 확인을 요청하는 용도이다.

    Args:
        wb_path: 입력 Excel 파일 경로
        config:  load_config()의 반환값

    Returns:
        누락 경고 목록. 모두 정상이면 빈 리스트.
    """
    import openpyxl  # noqa: PLC0415

    col_warnings: list[ColWarning] = []
    transaction_sheet = config.get("sheets", {}).get("transaction", "실제발생사업비")
    amount_cols: list[str] = config.get("amount_cols", [])
    nature_cols: list[str] = config.get("nature_cols", [])

    if not amount_cols and not nature_cols:
        return col_warnings

    try:
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    except Exception:
        return col_warnings

    try:
        if transaction_sheet not in wb.sheetnames:
            return col_warnings

        actual_cols = _get_sheet_columns(wb, transaction_sheet)
        actual_lower = {c.lower() for c in actual_cols}

        missing_amount = [c for c in amount_cols if c.lower() not in actual_lower]
        missing_nature = [c for c in nature_cols if c.lower() not in actual_lower]

        if missing_amount:
            col_warnings.append(ColWarning(
                sheet_key="transaction",
                category="amount_cols",
                missing_cols=missing_amount,
                found_cols=actual_cols,
            ))
        if missing_nature:
            col_warnings.append(ColWarning(
                sheet_key="transaction",
                category="nature_cols",
                missing_cols=missing_nature,
                found_cols=actual_cols,
            ))
    finally:
        wb.close()

    return col_warnings


# ════════════════════════════════════════════════════════════════════════════
#  Internal helpers
# ════════════════════════════════════════════════════════════════════════════

def _get_sheet_columns(wb, sheet_name: str, max_rows: int = 25) -> list[str]:
    """시트 상단 max_rows 행에서 유니크 문자열 값을 수집하여 반환한다.

    헤더 행 탐지 전략:
        - 가장 많은 문자열 셀을 가진 행을 헤더 행으로 간주하고 그 값을 앞에 배치한다.
        - 나머지 행의 값은 드롭다운 보조 후보로 뒤에 추가된다.
        - 순수 숫자 및 날짜 형식은 컬럼 후보에서 제외한다.

    출력 시트(출력>)처럼 sparse 구조인 경우에도 "출력전표" 레이블이
    상단 행에 존재하면 정상 탐지된다.
    """
    ws = wb[sheet_name]
    all_str_vals: set[str] = set()
    best_row_vals: list[str] = []
    best_count = 0

    for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
        if i >= max_rows:
            break
        str_vals = [
            str(v).strip()
            for v in row
            if v is not None
            and str(v).strip()
            and not _is_numeric_or_date(str(v).strip())
        ]
        all_str_vals.update(str_vals)
        if len(str_vals) > best_count:
            best_count = len(str_vals)
            best_row_vals = str_vals

    # 헤더 행 값 우선, 이후 나머지 보조 후보
    seen: set[str] = set()
    result: list[str] = []
    for v in best_row_vals + sorted(all_str_vals):
        if v not in seen:
            seen.add(v)
            result.append(v)
    return result


def _is_numeric_or_date(val: str) -> bool:
    """순수 숫자 또는 날짜 형식 문자열이면 True를 반환한다."""
    try:
        float(val.replace(",", ""))
        return True
    except ValueError:
        pass
    return bool(re.match(r"^\d{4}[-/\.]\d{2}[-/\.]\d{2}", val))


def _fuzzy_match(target: str, candidates: list[str], cutoff: float = 0.6) -> str | None:
    """difflib으로 target과 가장 유사한 후보를 반환한다. 없으면 None."""
    if not candidates or not target:
        return None
    matches = difflib.get_close_matches(target, candidates, n=1, cutoff=cutoff)
    return matches[0] if matches else None


def _deep_merge(base: dict, override: dict) -> dict:
    """base에 override를 재귀적으로 병합한다. override 값이 우선한다."""
    result = copy.deepcopy(base)
    for key, val in override.items():
        if key in result and isinstance(result[key], dict) and isinstance(val, dict):
            result[key] = _deep_merge(result[key], val)
        else:
            result[key] = val
    return result


# ════════════════════════════════════════════════════════════════════════════
#  YAML 파일 문자열 상수
# ════════════════════════════════════════════════════════════════════════════

_YAML_HEADER = """\
# =====================================================================
# 전표분석서 자동화 — 컬럼·시트명 설정 파일
# =====================================================================
# ⚠️  주의: Excel 파일의 컬럼명이나 시트명이 바뀌었을 때만 수정하세요.
#       수정 후 저장하면 다음 실행부터 새 값이 자동 적용됩니다.
# =====================================================================
# 형식: YAML (텍스트 파일, 메모장으로 열 수 있습니다)
#   각 항목:  논리명: "Excel에서 사용하는 실제 이름"
#   큰따옴표 안의 내용만 수정하고, 논리명(왼쪽)은 수정하지 마세요.
# =====================================================================

"""

_ANNOTATED_YAML_BODY = """\
sheets:
  # 각 시트의 실제 이름 (Excel 하단 탭 이름과 정확히 일치해야 합니다)
  transaction: "실제발생사업비"
  account: "계정정보"
  ccm: "Cost Center Master"
  output: "출력>"   # '>' 기호 포함 여부를 Excel 탭에서 직접 확인하세요

columns:
  # ── 실제발생사업비 시트 ─────────────────────────────────────────────────
  원가요소: "원가요소"
  코스트센터: "코스트 센터"             # 가운데 공백 포함

  # ── 계정정보 시트 ──────────────────────────────────────────────────────
  계정번호: "계정번호"
  계정명: "계정명"
  계정그룹ID: "계정그룹ID"
  계정그룹명: "계정그룹명"
  직간접구분_계정: "직/간접비"           # 슬래시(/) 포함
  대상정의: "대상정의 v3.0_0415"        # ⚠️  버전 번호 업데이트 시 여기를 수정
  범위: "사용 부서"
  지급대상: "비용 지급 범위"
  산출기준: "산출기준"

  # ── Cost Center Master 시트 ────────────────────────────────────────────
  cc_code: "Cost Center Code"
  cc_name: "Cost Center name"          # ⚠️  'n'이 소문자인지 대문자인지 확인
  직간접구분: "변경: 직접/공통 구분"

  # ── 출력전표 시트 ──────────────────────────────────────────────────────
  출력전표: "출력전표"
  귀속_주관부서: "귀속_주관부서"
  귀속_사용부서: "귀속_사용부서"

ccm_rename:
  # Cost Center Master 시트의 중복 열 이름 교체 규칙
  # pandas가 자동 생성하는 'Code.1' 형태를 의미 있는 이름으로 변환합니다.
  Code: "본부_Code"
  Code.1: "팀_Code"

amount_cols:
  # ⚠️  피벗 테이블의 금액 합계 컬럼명이 바뀌면 아래 목록을 수정하세요.
  # 누락된 항목은 해당 컬럼이 0으로 처리됩니다 (오류 없이 조용히 누락됨).
  - "Sum of DA_P"
  - "Sum of DA_N"
  - "Sum of DM"
  - "Sum of DC"
  - "Sum of DI"
  - "Sum of DP"
  - "Sum of IM"
  - "Sum of II"
  - "Sum of IO"
  - "Sum of IA"

nature_cols:
  # ⚠️  성격별 분류 컬럼명이 바뀌면 아래 목록을 수정하세요.
  # 순서도 보고서 테이블의 컬럼 순서에 영향을 줍니다.
  - "계약비"
  - "유지비"
  - "손해조사비"
  - "투자관리비"
  - "기타사업비"

col_team: "팀"   # 실제발생사업비 시트의 팀 컬럼명. 이름이 바뀌면 여기를 수정하세요.

output_labels:
  # 출력> 시트 안에서 위치 탐색에 사용하는 레이블 텍스트입니다.
  # Excel 담당자가 셀 레이블을 수정한 경우에만 변경하세요.
  코드: "출력전표"
  주관: "주관부서"
  사용: "사용부서"
"""
